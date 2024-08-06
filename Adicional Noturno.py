import os
from openpyxl import load_workbook
from datetime import datetime

current_datetime = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
str_current_datetime = str(current_datetime)

filepath = os.getenv('USERPROFILE')

workbook_hours = load_workbook(filepath + "\\Desktop\\" + "Gerador\\Adicional Noturno Professores.xlsx", data_only=True)

workbook_base = load_workbook(filepath + "\\Desktop\\" + "Gerador\\Planilha_Base_AN.xlsx")
worksheet_base = workbook_base ["Adicional Noturno"]

sheets = workbook_hours.sheetnames

if not os.path.exists(filepath + "\\Desktop\\" + "Gerador\\Gerados"):
    os.mkdir(filepath + "\\Desktop\\" + "Gerador\\Gerados")
else:
    print ("A pasta já existe")

count=0
list_addition=[]
unity_value=2

for worksheet_hours in workbook_hours.worksheets[0:1]:
    os.system('cls')
    print ("\nProcessando Informações da planilha", worksheet_hours.title,"\n")    
    for registration_cell, add_cell, add_min_cell, unity_cell in zip (worksheet_hours['A:A'],worksheet_hours['C:C'],worksheet_hours["E:E"], worksheet_hours["I:I"]): 
        if registration_cell.value is not None:
            if add_min_cell.value != 0 and add_min_cell.value is not None:
                if add_cell.value == "Sim":
                    registration = worksheet_hours.cell(row=registration_cell.row, column=1).value
                    name = worksheet_hours.cell(row=registration_cell.row, column=2).value
                    session = worksheet_hours.cell(row=registration_cell.row, column=4).value
                    time_convert = worksheet_hours.cell(row=registration_cell.row, column=7).value
                    decimal_minutes = ((time_convert * 60) / 100)
                    unity = worksheet_hours.cell(row=registration_cell.row, column=9).value
                    list_addition.append((registration, name, session, decimal_minutes, unity))
                    list_addition.sort(key=lambda list_addition: list_addition[4])
                    print (*list_addition, sep='\n')
                    worksheet_base.cell(row=count+5,column=1).value = list_addition[0+count][0]
                    worksheet_base.cell(row=count+5,column=2).value = list_addition[0+count][1]
                    worksheet_base.cell(row=count+5,column=3).value = list_addition[0+count][2]
                    worksheet_base.cell(row=count+5,column=7).value = list_addition[0+count][3]
                    worksheet_base.cell(row=count+5,column=11).value = list_addition[0+count][4]
                    workbook_base.save(filepath + "\\Desktop\\" + "Gerador\\Gerados\\Tabela Adicional Noturno Professores" + str(unity_value) + "(" + str_current_datetime + ")" + ".xlsx")
                    count += 1
#input("\nPressione qualquer tecla para encerrar a aplicação...")